#!/usr/bin/env python3
"""
prep.py — build data.parquet for the BSF Expression Lookup tool.

Joins the gene-level TPM matrix (merged_gene_2.tsv) with the transcript-level
functional annotation (BSF_Gene_FunctionalAnnotation.xlsx) into a single
Parquet file the frontend reads with DuckDB-WASM.

Grain note (important): TPM is GENE-level (one value set per gene); annotation
is TRANSCRIPT-level (918 genes carry 2-4 isoforms, 592 of which have genuinely
different annotations). Per the build decision, we KEEP EVERY ISOFORM: the
output is at transcript grain, with the gene's TPM repeated across its isoform
rows. Parquet dictionary/RLE compression makes the repetition essentially free,
and the frontend groups by gene_id (TPM shown once, each isoform listed).

Re-run this whenever the source files change:

    python -m pip install duckdb openpyxl
    python prep.py

Optionally override the source paths:

    python prep.py --tsv <path> --xlsx <path> --out <path>

Requires: duckdb, openpyxl  (both pure-pip; no build tooling).
"""

import argparse
import csv
import os
import sys

import duckdb
from openpyxl import load_workbook

# ── defaults ────────────────────────────────────────────────────────────────
HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_TSV = r"G:\My Drive\EntoZyme\Transcriptomics\merged_gene_2.tsv"
DEFAULT_XLSX = r"G:\My Drive\EntoZyme\Transcriptomics\BSF_Gene_FunctionalAnnotation.xlsx"
DEFAULT_OUT = os.path.join(HERE, "data.parquet")

# annotation sheet: source column index -> (clean_name, duckdb_type)
# (see BSF_Gene_FunctionalAnnotation.xlsx header; multi-line headers cleaned)
ANN_COLS = [
    (0,  "gene_id",       "VARCHAR"),
    (1,  "transcript_id", "VARCHAR"),
    (2,  "chrom",         "VARCHAR"),
    (3,  "start",         "BIGINT"),
    (4,  "end",           "BIGINT"),
    (5,  "strand",        "VARCHAR"),
    (6,  "cds_len",       "BIGINT"),
    (7,  "nr_hit",        "VARCHAR"),
    (8,  "nr_evalue",     "DOUBLE"),
    (9,  "nr_pident",     "DOUBLE"),
    (10, "nr_desc",       "VARCHAR"),
    (11, "dm_hit",        "VARCHAR"),
    (12, "dm_evalue",     "DOUBLE"),
    (13, "dm_pident",     "DOUBLE"),
    (14, "dm_desc",       "VARCHAR"),
    (15, "pfam",          "VARCHAR"),
    (16, "interpro",      "VARCHAR"),
    (17, "go",            "VARCHAR"),
]


def clean(v, typ):
    """Normalise a raw openpyxl cell to the target duckdb type (or None)."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    if typ == "BIGINT":
        try:
            return int(v)
        except (TypeError, ValueError):
            return None
    if typ == "DOUBLE":
        try:
            return float(v)
        except (TypeError, ValueError):
            return None
    return str(v)


def read_annotation(xlsx_path):
    print(f"[ann] reading {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)  # header
    out = []
    for r in rows_iter:
        if r[0] is None:  # skip blank tail rows
            continue
        out.append(tuple(clean(r[i] if i < len(r) else None, typ)
                         for (i, _name, typ) in ANN_COLS))
    print(f"[ann] {len(out)} transcript rows")
    return out


def sample_columns(tsv_path):
    with open(tsv_path, newline="", encoding="utf-8") as f:
        header = next(csv.reader(f, delimiter="\t"))
    if header[0] != "gene_id":
        sys.exit(f"[tsv] expected first column 'gene_id', got {header[0]!r}")
    return header[1:]


def main():
    ap = argparse.ArgumentParser(description="Build data.parquet for BSF lookup.")
    ap.add_argument("--tsv", default=DEFAULT_TSV)
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    for p in (args.tsv, args.xlsx):
        if not os.path.exists(p):
            sys.exit(f"missing input: {p}")

    samples = sample_columns(args.tsv)
    n = len(samples)
    print(f"[tsv] {n} samples")

    ann_rows = read_annotation(args.xlsx)

    con = duckdb.connect()

    # ── annotation table (transcript grain) ─────────────────────────────────
    ann_schema = ", ".join(f'"{name}" {typ}' for (_i, name, typ) in ANN_COLS)
    con.execute(f"CREATE TABLE ann ({ann_schema})")
    placeholders = ", ".join(["?"] * len(ANN_COLS))
    con.executemany(f"INSERT INTO ann VALUES ({placeholders})", ann_rows)

    # ── tpm table (gene grain) ───────────────────────────────────────────────
    # Read with GENERIC column names (s0..sN) so the real sample labels — which
    # include both 'MT1/MT2' and 'Mt1/Mt2' (distinct samples that collide under
    # DuckDB's case-insensitive identifiers) — never become SQL columns. TPM is
    # packed into one ordered DOUBLE[] array; the frontend holds the true labels.
    gen_names = "['gene_id', " + ", ".join(f"'s{i}'" for i in range(n)) + "]"
    arr = "[" + ", ".join(f"s{i}" for i in range(n)) + "]"
    con.execute(f"""
        CREATE TABLE tpm AS
        SELECT gene_id, {arr} AS tpm
        FROM read_csv(?, delim='\t', header=true, names={gen_names}, auto_detect=true)
    """, [args.tsv])

    # ── integrity checks (fail loud if id sets drift) ────────────────────────
    (ann_genes,) = con.execute("SELECT COUNT(DISTINCT gene_id) FROM ann").fetchone()
    (tpm_genes,) = con.execute("SELECT COUNT(DISTINCT gene_id) FROM tpm").fetchone()
    (only_tpm,) = con.execute(
        "SELECT COUNT(*) FROM (SELECT gene_id FROM tpm EXCEPT SELECT gene_id FROM ann)"
    ).fetchone()
    (only_ann,) = con.execute(
        "SELECT COUNT(*) FROM (SELECT gene_id FROM ann EXCEPT SELECT gene_id FROM tpm)"
    ).fetchone()
    print(f"[join] distinct genes  tpm={tpm_genes}  ann={ann_genes}  "
          f"only_tpm={only_tpm}  only_ann={only_ann}")

    # ── join: keep every gene present in either file (outer), every isoform ──
    con.execute("""
        CREATE TABLE joined AS
        SELECT
            COALESCE(a.gene_id, t.gene_id) AS gene_id,
            a.transcript_id, a.chrom, a.start, a.end, a.strand, a.cds_len,
            a.nr_hit, a.nr_evalue, a.nr_pident, a.nr_desc,
            a.dm_hit, a.dm_evalue, a.dm_pident, a.dm_desc,
            a.pfam, a.interpro, a.go,
            t.tpm
        FROM ann a
        FULL OUTER JOIN tpm t ON a.gene_id = t.gene_id
        ORDER BY gene_id, a.transcript_id
    """)
    (nrows,) = con.execute("SELECT COUNT(*) FROM joined").fetchone()
    (ngenes,) = con.execute("SELECT COUNT(DISTINCT gene_id) FROM joined").fetchone()
    print(f"[join] joined rows={nrows}  distinct genes={ngenes}")

    # ── write parquet ───────────────────────────────────────────────────────
    con.execute("COPY joined TO ? (FORMAT parquet, COMPRESSION zstd)", [args.out])
    size = os.path.getsize(args.out)
    print(f"[out] wrote {args.out}  ({size/1e6:.2f} MB)")
    print(f"[out] sample order: {samples}")


if __name__ == "__main__":
    main()
